import argparse
import requests
import openpyxl
import json
import time
import re
import ast
import urllib3
from urllib.parse import urljoin
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, PatternFill, Border, Side

HTTP_METHODS = {"GET", "POST", "PATCH", "PUT", "DELETE"}
MACRO_PATTERN = re.compile(r'\$\{([^}]+)\}')

# ignore SSL warnings
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

def delay_function(endpoint):
    try:
        delay_seconds = int(endpoint)
        print(f"[*] Delay for {delay_seconds} seconds")
        for remaining in range(delay_seconds, 0, -1):
            print(f"    Remaining: {remaining} seconds", end='\r')
            time.sleep(1)
        print("    Delay complete. Moving to next command.\n")
    except ValueError:
        print(f"[!] Invalid delay time: {endpoint}")

def parse_change_username_endpoint(endpoint, username_to_id_map):
    pattern = r'\${([^.]+)\.id}'
    match = re.search(pattern, endpoint)
    if match:
        username_key = match.group(1)
        if username_key in username_to_id_map:
            endpoint = re.sub(pattern, username_to_id_map[username_key], endpoint)
            print(f"username_key: {username_key}, ID: {username_to_id_map[username_key]}")
            print(f"[*] Replaced dynamic endpoint with ID: {endpoint}")
        else:
            print(f"[!] Warning: No ID found for username {username_key}")
    return endpoint

def get_nested_value(data, path_parts):
    current = data
    for part in path_parts:
        if not isinstance(current, dict) or part not in current:
            raise KeyError(f"Path not found: {'.'.join(path_parts)}")
        current = current[part]
    return current

def set_nested_value(data, path_parts, value):
    current = data
    for part in path_parts[:-1]:
        if part not in current or not isinstance(current[part], dict):
            current[part] = {}
        current = current[part]
    current[path_parts[-1]] = value

def parse_literal_value(value):
    if not isinstance(value, str):
        return value

    stripped = value.strip()
    if stripped == "":
        return ""

    try:
        return json.loads(stripped)
    except json.JSONDecodeError:
        return stripped

def resolve_macro_expr(expr, global_context, batch_context, response_json, status_code, username_to_id_map):
    if expr == "STATUSCODE":
        if status_code is None:
            raise KeyError("STATUSCODE is not available before HTTP execution")
        return status_code

    if expr.startswith("RESPONSE."):
        if response_json is None:
            raise KeyError("RESPONSE is not available before HTTP execution")
        return get_nested_value(response_json, expr.split(".")[1:])

    if expr.startswith("CONTEXT."):
        return get_nested_value(global_context, expr.split(".")[1:])

    if expr.startswith("BATCH.CONTEXT."):
        if batch_context is None:
            raise KeyError("BATCH.CONTEXT is only available during batch execution")
        return get_nested_value(batch_context, expr.split(".")[2:])

    # Backward compatibility: ${username.id}
    legacy_match = re.fullmatch(r'([^.]+)\.id', expr)
    if legacy_match:
        username_key = legacy_match.group(1)
        if username_key in username_to_id_map:
            return username_to_id_map[username_key]
        raise KeyError(f"No ID found for username: {username_key}")

    # NAME namespace compatibility: ${NAME.username.id}
    name_match = re.fullmatch(r'NAME\.([^.]+)\.id', expr)
    if name_match:
        username_key = name_match.group(1)
        if username_key in username_to_id_map:
            return username_to_id_map[username_key]
        raise KeyError(f"No ID found for NAME username: {username_key}")

    raise KeyError(f"Unsupported macro: ${{{expr}}}")

def substitute_text(raw_value, global_context, batch_context, response_json, status_code, username_to_id_map):
    if raw_value is None:
        return None
    if not isinstance(raw_value, str):
        return raw_value

    text = raw_value.strip()
    if text == "":
        return ""

    full_match = re.fullmatch(r'\$\{([^}]+)\}', text)
    if full_match:
        return resolve_macro_expr(
            full_match.group(1),
            global_context,
            batch_context,
            response_json,
            status_code,
            username_to_id_map
        )

    def replacer(match):
        value = resolve_macro_expr(
            match.group(1),
            global_context,
            batch_context,
            response_json,
            status_code,
            username_to_id_map
        )
        if isinstance(value, (dict, list)):
            return json.dumps(value, ensure_ascii=False)
        return str(value)

    return MACRO_PATTERN.sub(replacer, text)

def parse_if_condition(condition_expr, global_context, batch_context, response_json, status_code, username_to_id_map):
    def to_expr_literal(value):
        if isinstance(value, bool):
            return "True" if value else "False"
        if value is None:
            return "None"
        if isinstance(value, (int, float)):
            return str(value)
        if isinstance(value, (dict, list)):
            return json.dumps(value, ensure_ascii=False)
        return json.dumps(str(value), ensure_ascii=False)

    def substitute_condition_expression(expr):
        def replacer(match):
            resolved = resolve_macro_expr(
                match.group(1),
                global_context,
                batch_context,
                response_json,
                status_code,
                username_to_id_map
            )
            return to_expr_literal(resolved)

        substituted = MACRO_PATTERN.sub(replacer, expr)

        # Support lowercase JSON-style booleans/null in conditions.
        substituted = re.sub(r'\btrue\b', 'True', substituted, flags=re.IGNORECASE)
        substituted = re.sub(r'\bfalse\b', 'False', substituted, flags=re.IGNORECASE)
        substituted = re.sub(r'\bnull\b', 'None', substituted, flags=re.IGNORECASE)
        return substituted

    def eval_compare_operator(op_node, left, right):
        if isinstance(op_node, ast.Eq):
            return left == right
        if isinstance(op_node, ast.NotEq):
            return left != right
        if isinstance(op_node, ast.Gt):
            return left > right
        if isinstance(op_node, ast.Lt):
            return left < right
        if isinstance(op_node, ast.GtE):
            return left >= right
        if isinstance(op_node, ast.LtE):
            return left <= right
        raise ValueError("Unsupported comparison operator in IF condition")

    def eval_node(node):
        if isinstance(node, ast.Constant):
            return node.value

        if isinstance(node, ast.Name):
            if node.id in {"True", "False", "None"}:
                return {"True": True, "False": False, "None": None}[node.id]
            raise ValueError(f"Unsupported identifier in IF condition: {node.id}")

        if isinstance(node, ast.UnaryOp) and isinstance(node.op, ast.Not):
            return not bool(eval_node(node.operand))

        if isinstance(node, ast.BoolOp):
            if isinstance(node.op, ast.And):
                result = True
                for value_node in node.values:
                    result = result and bool(eval_node(value_node))
                    if not result:
                        break
                return result
            if isinstance(node.op, ast.Or):
                result = False
                for value_node in node.values:
                    result = result or bool(eval_node(value_node))
                    if result:
                        break
                return result
            raise ValueError("Unsupported boolean operator in IF condition")

        if isinstance(node, ast.Compare):
            left = eval_node(node.left)
            for op_node, comparator_node in zip(node.ops, node.comparators):
                right = eval_node(comparator_node)
                if not eval_compare_operator(op_node, left, right):
                    return False
                left = right
            return True

        if isinstance(node, ast.Expression):
            return eval_node(node.body)

        raise ValueError("Unsupported expression in IF condition")

    substituted_expr = substitute_condition_expression(condition_expr.strip())
    try:
        tree = ast.parse(substituted_expr, mode='eval')
    except SyntaxError as exc:
        raise ValueError(f"Invalid IF condition syntax: {condition_expr}") from exc

    return bool(eval_node(tree))

def extract_batch_signature(method_value):
    match = re.fullmatch(r'BATCH_START\(([^)]+)\)', method_value.strip(), re.IGNORECASE)
    if match:
        return "start", match.group(1).strip(), None

    match = re.fullmatch(r'BATCH_END\(([^)]+)\)', method_value.strip(), re.IGNORECASE)
    if match:
        return "end", match.group(1).strip(), None

    match = re.fullmatch(r'BATCH_END', method_value.strip(), re.IGNORECASE)
    if match:
        return "end", None, None

    match = re.fullmatch(r'BATCH\(([^,]+),(.*)\)', method_value.strip(), re.IGNORECASE)
    if match:
        return "run", match.group(1).strip(), match.group(2).strip()

    match = re.fullmatch(r'BATCH\(([^)]+)\)', method_value.strip(), re.IGNORECASE)
    if match:
        return "run", match.group(1).strip(), None

    return None, None, None

def parse_commands(sheet):
    commands = []
    for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), 2):
        cells = list(row) if row is not None else []
        while len(cells) < 4:
            cells.append(None)

        method, endpoint, payload, request_rule = cells[:4]
        if method is None or str(method).strip() == "":
            continue

        commands.append({
            "row_num": row_num,
            "method": str(method).strip(),
            "endpoint": "" if endpoint is None else str(endpoint),
            "payload": None if payload is None else str(payload),
            "request": "" if request_rule is None else str(request_rule)
        })
    return commands

def apply_request_rules(request_text, global_context, batch_context, response_json, status_code, username_to_id_map):
    if request_text is None or request_text.strip() == "":
        return

    for raw_line in request_text.splitlines():
        line = raw_line.strip()
        if line == "":
            continue
        if "=" not in line:
            raise ValueError(f"Invalid Request syntax: {line}")

        left, right = line.split("=", 1)
        left = left.strip()
        right = right.strip()

        target_match = re.fullmatch(r'\$\{([^}]+)\}', left)
        if not target_match:
            raise ValueError(f"Invalid Request target: {left}")

        target_expr = target_match.group(1)
        right_sub = substitute_text(
            right,
            global_context,
            batch_context,
            response_json,
            status_code,
            username_to_id_map
        )
        value = parse_literal_value(right_sub)

        if target_expr.startswith("CONTEXT."):
            set_nested_value(global_context, target_expr.split(".")[1:], value)
        elif target_expr.startswith("BATCH.CONTEXT."):
            if batch_context is None:
                raise ValueError("BATCH.CONTEXT cannot be written outside batch execution")
            set_nested_value(batch_context, target_expr.split(".")[2:], value)
        else:
            raise ValueError(f"Unsupported Request target: {left}")

def update_username_id(method, endpoint, status_code, response_json, username_to_id_map):
    # Store username to ID mapping for account creation
    check_then_store = False
    if (method.upper() == "POST" and 
        endpoint.startswith("/redfish/v1/AccountService/Accounts") and 
        status_code == 201):
        check_then_store = True
        
    # Also store username to ID mapping from GET requests to specific account endpoints
    elif (method.upper() == "GET" and 
            endpoint.startswith("/redfish/v1/AccountService/Accounts") and
            endpoint != "/redfish/v1/AccountService/Accounts/" and
            status_code == 200):
        check_then_store = True

    # Also store username to ID mapping from update operations (PATCH/PUT)
    elif ((method.upper() == "PATCH" or method.upper() == "PUT") and 
            endpoint.startswith("/redfish/v1/AccountService/Accounts") and
            endpoint != "/redfish/v1/AccountService/Accounts/" and
            status_code in [200, 202, 204]):
        check_then_store = True
 
    if check_then_store:
        if "UserName" in response_json and "Id" in response_json:
            username_value = response_json["UserName"]
            id_value = response_json["Id"]
            username_to_id_map[username_value] = id_value
            print(f"[*] Stored mapping: {username_value} -> {id_value}")

def get_id_username_map(root_url,username, password, username_to_id_map):
    # Add get method with /redfish/v1/AccountService/Accounts endpoint to fetch existing accounts
    get_accounts_endpoint = "/redfish/v1/AccountService/Accounts"
    print(f"[*] Fetching existing accounts from {get_accounts_endpoint}")
    url = urljoin(root_url, get_accounts_endpoint)
    headers = {"Content-Type": "application/json"}
    auth = (username, password)
    try:
        response = requests.get(
            url=url,
            auth=auth,
            headers=headers,
            verify=False,
            timeout=10
        )
        if response.status_code == 200:
            response_json = response.json()
            if isinstance(response_json, dict) and "Members" in response_json:
                for member in response_json["Members"]:
                    #if "@odata.id" in member, use it as endpoint to fetch account details
                    if "@odata.id" in member:
                        account_url = urljoin(root_url, member["@odata.id"])
                        account_response = requests.get(
                            url=account_url,
                            auth=auth,
                            headers=headers,
                            verify=False,
                            timeout=10
                        )
                        if account_response.status_code == 200:
                            account_data = account_response.json()
                            if "UserName" in account_data and "Id" in account_data:
                                username_to_id_map[account_data["UserName"]] = account_data["Id"]
                                print(f"[*] Found existing account: {account_data['UserName']} -> {account_data['Id']}")
                        else:
                            print(f"[!] Failed to fetch account details: {account_response.status_code} - {account_response.text}")
                    if "UserName" in member and "Id" in member:
                        username_to_id_map[member["UserName"]] = member["Id"]
                        print(f"[*] Found existing account: {member['UserName']} -> {member['Id']}")
            print(f"[*] Finished fetching existing accounts.")
        else:
            print(f"[!] Failed to fetch accounts: {response.status_code} - {response.text}")
    except Exception as e:
        print(f"[!] Error fetching accounts: {e}")

BATCH_ROW_FILL = PatternFill(fill_type="solid", start_color="FFFDE9D9", end_color="FFFDE9D9")
BATCH_ROW_BORDER = Border(
    left=Side(style="thin", color="FFD9D9D9"),
    right=Side(style="thin", color="FFD9D9D9"),
    top=Side(style="thin", color="FFD9D9D9"),
    bottom=Side(style="thin", color="FFD9D9D9")
)


def append_output_row(output_sheet, output_row_index, row_data, highlight_batch=False):
    if isinstance(output_sheet, list):
        output_sheet.append(row_data)
        return

    output_sheet.append(row_data)

    for col_num, cell_value in enumerate(row_data, 1):
        column_letter = get_column_letter(col_num)
        current_width = output_sheet.column_dimensions[column_letter].width
        if current_width is None:
            current_width = 10
        output_sheet.column_dimensions[column_letter].width = max(current_width, len(str(cell_value)) + 2)

        cell = output_sheet.cell(row=output_row_index, column=col_num)
        cell.alignment = Alignment(wrap_text=True)
        if highlight_batch:
            cell.fill = BATCH_ROW_FILL
            cell.border = BATCH_ROW_BORDER

    response_text = str(row_data[-1]) if row_data else ""
    num_lines = response_text.count('\n') + 1
    output_sheet.row_dimensions[output_row_index].height = max(15, 15 * num_lines)

def execute_command_block(
    commands,
    username,
    password,
    root_url,
    output_sheet,
    output_row_index,
    global_context,
    username_to_id_map,
    batches,
    in_batch=False,
    batch_context=None,
    batch_invocation_label=None
):
    headers = {"Content-Type": "application/json"}
    auth = (username, password)
    if_stack = []

    def current_active():
        return all(frame["current_active"] for frame in if_stack)

    index = 0
    while index < len(commands):
        command = commands[index]
        method_raw = command["method"].strip()
        method_upper = method_raw.upper()
        endpoint_raw = command["endpoint"]
        payload_raw = command["payload"]
        request_raw = command["request"]

        signature_type, signature_name, signature_arg = extract_batch_signature(method_raw)

        if signature_type == "start":
            if in_batch:
                append_output_row(
                    output_sheet,
                    output_row_index,
                    [method_raw, endpoint_raw, payload_raw, request_raw, "Error", "Nested BATCH definition is not supported"]
                )
                output_row_index += 1
                return {"output_row_index": output_row_index, "terminated": False, "batch_status": "ERROR"}

            batch_name = signature_name
            block = []
            cursor = index + 1
            found_end = False
            while cursor < len(commands):
                next_cmd = commands[cursor]
                next_sig_type, next_sig_name, _ = extract_batch_signature(next_cmd["method"])
                if next_sig_type == "end" and (next_sig_name is None or next_sig_name == batch_name):
                    found_end = True
                    break
                block.append(next_cmd)
                cursor += 1

            if not found_end:
                append_output_row(
                    output_sheet,
                    output_row_index,
                    [method_raw, endpoint_raw, payload_raw, request_raw, "Error", f"BATCH_END not found for {batch_name}"]
                )
                output_row_index += 1
                return {"output_row_index": output_row_index, "terminated": False, "batch_status": "ERROR"}

            batches[batch_name] = {
                "commands": block,
                "input_schema": payload_raw or "",
            }
            print(f"[*] Registered batch: {batch_name} ({len(block)} commands)")
            index = cursor + 1
            continue

        if signature_type == "end":
            append_output_row(
                output_sheet,
                output_row_index,
                [method_raw, endpoint_raw, payload_raw, request_raw, "Error", "BATCH_END without matching BATCH_START"]
            )
            output_row_index += 1
            index += 1
            continue

        if method_upper.startswith("IF ") or method_upper == "IF":
            parent_active = current_active()
            if payload_raw is not None and str(payload_raw).strip() != "":
                condition_expr = str(payload_raw).strip()
            elif method_upper.startswith("IF "):
                # Backward compatibility: IF condition inline in Method.
                condition_expr = method_raw[3:].strip()
            else:
                # Backward compatibility: IF condition in Endpoint.
                condition_expr = endpoint_raw.strip()
            condition_true = False
            error_message = None

            if parent_active:
                try:
                    condition_true = parse_if_condition(
                        condition_expr,
                        global_context,
                        batch_context,
                        None,
                        None,
                        username_to_id_map
                    )
                except Exception as exc:
                    error_message = str(exc)

            frame = {
                "parent_active": parent_active,
                "condition_true": condition_true,
                "current_active": parent_active and condition_true,
                "else_seen": False
            }
            if_stack.append(frame)

            if error_message:
                append_output_row(
                    output_sheet,
                    output_row_index,
                    [method_raw, endpoint_raw, payload_raw, request_raw, "Error", f"IF parse error: {error_message}"]
                )
                output_row_index += 1

            index += 1
            continue

        if method_upper == "ELSE":
            if not if_stack:
                append_output_row(
                    output_sheet,
                    output_row_index,
                    [method_raw, endpoint_raw, payload_raw, request_raw, "Error", "ELSE without IF"]
                )
                output_row_index += 1
                index += 1
                continue

            top = if_stack[-1]
            if top["else_seen"]:
                append_output_row(
                    output_sheet,
                    output_row_index,
                    [method_raw, endpoint_raw, payload_raw, request_raw, "Error", "Duplicate ELSE in IF block"]
                )
                output_row_index += 1
            else:
                top["else_seen"] = True
                top["current_active"] = top["parent_active"] and (not top["condition_true"])

            index += 1
            continue

        if method_upper == "ENDIF":
            if not if_stack:
                append_output_row(
                    output_sheet,
                    output_row_index,
                    [method_raw, endpoint_raw, payload_raw, request_raw, "Error", "ENDIF without IF"]
                )
                output_row_index += 1
            else:
                if_stack.pop()
            index += 1
            continue

        if not current_active():
            index += 1
            continue

        if signature_type == "run":
            batch_name = signature_name
            batch_method_label = f"BATCH({batch_name})"
            if payload_raw is not None and str(payload_raw).strip() != "":
                argument_raw = str(payload_raw).strip()
            else:
                # Backward compatibility: BATCH argument inline in Method.
                argument_raw = signature_arg
            if batch_name not in batches:
                append_output_row(
                    output_sheet,
                    output_row_index,
                    [method_raw, endpoint_raw, payload_raw, request_raw, "Error", f"Batch not found: {batch_name}"]
                )
                output_row_index += 1
                index += 1
                continue

            batch_def = batches[batch_name]
            if isinstance(batch_def, dict):
                batch_commands = batch_def.get("commands", [])
                batch_input_schema = batch_def.get("input_schema", "")
            else:
                batch_commands = batch_def
                batch_input_schema = ""

            try:
                if argument_raw in [None, ""]:
                    argument_value = {}
                else:
                    argument_sub = substitute_text(
                        argument_raw,
                        global_context,
                        batch_context,
                        None,
                        None,
                        username_to_id_map
                    )
                    argument_value = parse_literal_value(argument_sub)
            except Exception as exc:
                append_output_row(
                    output_sheet,
                    output_row_index,
                    [method_raw, endpoint_raw, payload_raw, request_raw, "Error", f"Batch argument error: {exc}"]
                )
                output_row_index += 1
                index += 1
                continue

            local_batch_context = {}
            local_batch_context["input"] = argument_value
            if isinstance(argument_value, dict):
                for key, value in argument_value.items():
                    local_batch_context[key] = value

            batch_output_rows = []

            result = execute_command_block(
                batch_commands,
                username,
                password,
                root_url,
                batch_output_rows,
                1,
                global_context,
                username_to_id_map,
                batches,
                in_batch=True,
                batch_context=local_batch_context,
                batch_invocation_label=batch_method_label
            )

            summary_row = result.get("summary_row")
            if summary_row is None:
                response_text = json.dumps(local_batch_context, indent=4, ensure_ascii=False)
                summary_row = [batch_method_label, endpoint_raw, payload_raw, request_raw, result["batch_status"] or "END", response_text]

            append_output_row(
                output_sheet,
                output_row_index,
                summary_row
            )
            output_row_index += 1

            append_output_row(
                output_sheet,
                output_row_index,
                [f"BATCH_START({batch_name})", endpoint_raw, batch_input_schema, request_raw, "START", ""]
                ,
                highlight_batch=True
            )
            output_row_index += 1

            for batch_row in batch_output_rows:
                append_output_row(output_sheet, output_row_index, batch_row, highlight_batch=True)
                output_row_index += 1

            append_output_row(
                output_sheet,
                output_row_index,
                [f"BATCH_END({batch_name})", endpoint_raw, "", request_raw, result["batch_status"] or "END", ""]
                ,
                highlight_batch=True
            )
            output_row_index += 1
            index += 1
            continue

        if method_upper in {"SUCCESS", "ERROR"}:
            if not in_batch:
                append_output_row(
                    output_sheet,
                    output_row_index,
                    [method_raw, endpoint_raw, payload_raw, request_raw, "Error", f"{method_upper} can only be used in batch"]
                )
                output_row_index += 1
                index += 1
                continue

            response_text = json.dumps(batch_context if batch_context is not None else {}, indent=4, ensure_ascii=False)
            summary_row = [batch_invocation_label or method_upper, endpoint_raw, payload_raw, request_raw, method_upper, response_text]
            if isinstance(output_sheet, list):
                return {
                    "output_row_index": output_row_index,
                    "terminated": True,
                    "batch_status": method_upper,
                    "summary_row": summary_row
                }

            append_output_row(
                output_sheet,
                output_row_index,
                summary_row
            )
            output_row_index += 1
            return {"output_row_index": output_row_index, "terminated": True, "batch_status": method_upper, "summary_row": summary_row}

        if method_upper == "BATCH_DESC":
            try:
                desc_value = substitute_text(
                    payload_raw,
                    global_context,
                    batch_context,
                    None,
                    None,
                    username_to_id_map
                )
                if isinstance(desc_value, (dict, list)):
                    response_text = json.dumps(desc_value, indent=4, ensure_ascii=False)
                elif desc_value is None:
                    response_text = ""
                else:
                    response_text = str(desc_value)

                append_output_row(
                    output_sheet,
                    output_row_index,
                    [method_raw, endpoint_raw, payload_raw, request_raw, "BATCH_DESC", response_text]
                )
            except Exception as exc:
                append_output_row(
                    output_sheet,
                    output_row_index,
                    [method_raw, endpoint_raw, payload_raw, request_raw, "Error", f"BATCH_DESC substitution error: {exc}"]
                )
            output_row_index += 1
            index += 1
            continue

        if method_upper == "MESSAGE":
            try:
                message_value = substitute_text(
                    payload_raw,
                    global_context,
                    batch_context,
                    None,
                    None,
                    username_to_id_map
                )
                if isinstance(message_value, (dict, list)):
                    response_text = json.dumps(message_value, indent=4, ensure_ascii=False)
                elif message_value is None:
                    response_text = ""
                else:
                    response_text = str(message_value)

                append_output_row(
                    output_sheet,
                    output_row_index,
                    [method_raw, endpoint_raw, payload_raw, request_raw, "MESSAGE", response_text]
                )
            except Exception as exc:
                append_output_row(
                    output_sheet,
                    output_row_index,
                    [method_raw, endpoint_raw, payload_raw, request_raw, "Error", f"MESSAGE substitution error: {exc}"]
                )
            output_row_index += 1
            index += 1
            continue

        if method_upper == "DELAY":
            delay_function(endpoint_raw)
            index += 1
            continue

        if method_upper not in HTTP_METHODS:
            append_output_row(
                output_sheet,
                output_row_index,
                [method_raw, endpoint_raw, payload_raw, request_raw, "Error", f"Unsupported method: {method_raw}"]
            )
            output_row_index += 1
            index += 1
            continue

        status_code = None
        response_json = None
        response_text = ""
        
        #print current username_to_id_map before substitution to help debugging
        print(f"[*] Executing command at row {command['row_num']}: {method_upper} {endpoint_raw}")
        print(f"[*] Current username_to_id_map before substitution: {json.dumps(username_to_id_map, ensure_ascii=False)}")

        try:
            endpoint = substitute_text(
                endpoint_raw,
                global_context,
                batch_context,
                None,
                None,
                username_to_id_map
            )
            if not isinstance(endpoint, str):
                endpoint = str(endpoint)

            payload_sub = substitute_text(
                payload_raw,
                global_context,
                batch_context,
                None,
                None,
                username_to_id_map
            )

            data = None
            if payload_sub not in [None, ""]:
                if isinstance(payload_sub, (dict, list)):
                    data = payload_sub
                else:
                    data = json.loads(str(payload_sub))

            url = urljoin(root_url, endpoint)
            print(f"[*] {method_upper} {url}")
            response = requests.request(
                method=method_upper,
                url=url,
                auth=auth,
                headers=headers,
                json=data,
                verify=False,
                timeout=10
            )
            status_code = response.status_code

            try:
                response_json = response.json()
                response_text = json.dumps(response_json, indent=4, ensure_ascii=False)
                update_username_id(method_upper, endpoint, status_code, response_json, username_to_id_map)
            except json.JSONDecodeError:
                response_text = response.text

            is_create_account = (
                method_upper == "POST"
                and endpoint == "/redfish/v1/AccountService/Accounts"
            )
            is_delete_account = (
                method_upper == "DELETE"
                and endpoint.startswith("/redfish/v1/AccountService/Accounts/")
                and endpoint != "/redfish/v1/AccountService/Accounts/"
            )
            if is_create_account or is_delete_account:
                print(f"[*] username_to_id_map after {method_upper} {endpoint}: {json.dumps(username_to_id_map, ensure_ascii=False)}")

            request_resolved_text = request_raw
            if request_raw and request_raw.strip() != "":
                apply_request_rules(
                    request_raw,
                    global_context,
                    batch_context,
                    response_json,
                    status_code,
                    username_to_id_map
                )

            append_output_row(
                output_sheet,
                output_row_index,
                [method_upper, endpoint, payload_raw, request_resolved_text, status_code, response_text]
            )
            output_row_index += 1

            print(f"Status Code: {status_code}")
            print(f"Response: {response_text}\n")
            
            #print current username_to_id_map before substitution to help debugging
            print(f"[*] After command at row {command['row_num']}: {method_upper} {endpoint_raw}")
            print(f"[*] Current username_to_id_map after substitution: {json.dumps(username_to_id_map, ensure_ascii=False)}")

        except Exception as exc:
            append_output_row(
                output_sheet,
                output_row_index,
                [method_upper, endpoint_raw, payload_raw, request_raw, "Error", str(exc)]
            )
            output_row_index += 1
            print(f"[!] request or processing error: {exc}")

        index += 1

    if if_stack:
        append_output_row(
            output_sheet,
            output_row_index,
            ["IF", "", "", "", "Error", "Unclosed IF block (missing ENDIF)"]
        )
        output_row_index += 1

    return {"output_row_index": output_row_index, "terminated": False, "batch_status": None, "summary_row": None}

def execute_redfish(username, password, root_url, excel_path='commands.xlsx', output_excel_path='output.xlsx'):
    try:
        wb = openpyxl.load_workbook(excel_path)
        sheet = wb.active

        # build Excel file for output
        output_wb = openpyxl.Workbook()
        output_sheet = output_wb.active
        if output_sheet is None:
            raise RuntimeError("Failed to get active worksheet for output")
        output_sheet.append(["Method", "Endpoint", "Payload", "Request", "Status Code", "Response"])  # write title

        # Dictionary to store username to id mappings
        username_to_id_map = {}

        get_id_username_map(root_url,username, password, username_to_id_map)
        print(f"[*] Current username to ID mapping: {username_to_id_map}")

        for col_num, column_title in enumerate(["Method", "Endpoint", "Payload", "Request", "Status Code", "Response"], 1):
            column_letter = get_column_letter(col_num)
            output_sheet.column_dimensions[column_letter].width = max(10, len(column_title) + 2)  # Minimum width of 10

        commands = parse_commands(sheet)
        batches = {}
        global_context = {}

        result = execute_command_block(
            commands,
            username,
            password,
            root_url,
            output_sheet,
            output_row_index=2,
            global_context=global_context,
            username_to_id_map=username_to_id_map,
            batches=batches,
            in_batch=False,
            batch_context=None
        )

        print(f"[*] Final CONTEXT: {json.dumps(global_context, ensure_ascii=False)}")
        print(f"[*] Batch definitions: {list(batches.keys())}")
        print(f"[*] Output rows written: {result['output_row_index'] - 2}")

        # 儲存輸出 Excel 檔案
        output_wb.save(output_excel_path)
        print(f"[*] Output result to：{output_excel_path}")

    except FileNotFoundError:
        print(f"[!] Excel file not found: {excel_path}")
    except Exception as e:
        print(f"[!] Execution error: {e}")

def main():
    parser = argparse.ArgumentParser(description="Execute Redfish commands from Excel.")
    parser.add_argument("-u", "--username", required=True, help="Username")
    parser.add_argument("-p", "--password", required=True, help="Password")
    parser.add_argument("-r", "--root", required=True, help="Root URL (e.g., https://127.0.0.1:5101)")
    parser.add_argument("-f", "--file", default="commands.xlsx", help="Excel file path")
    parser.add_argument("-o", "--output", default="output.xlsx", help="Output Excel file path") 

    args = parser.parse_args()

    execute_redfish(args.username, args.password, args.root, args.file, args.output)  

if __name__ == "__main__":
    main()

